import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import json
import re
import xlwings as xw
from excel_manager import ExcelManager

EXCEL_FILE = 'Container đã nhập tháng 5 - Minh Quang.xlsx'
SETTINGS_FILE = 'settings.json'

FIELDS = [
    'NGÀY LẤY',
    'CTY',
    'NHÀ XE',
    'BK No',
    'MÃ SỐ CONTAINER',
    'SEAL',
    'Loại hình',
    'Số lượng',
    'Kích cỡ',
    'NƠI LẤY CONT',
    'NƠI HẠ CONT'
]

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"sheets": {}}

def save_settings(settings):
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def col_letter_to_index(col):
    # Chuyển chữ cái cột Excel (A, B, AA, ...) thành số thứ tự (1, 2, 27, ...)
    col = col.upper()
    num = 0
    for c in col:
        if 'A' <= c <= 'Z':
            num = num * 26 + (ord(c) - ord('A') + 1)
    return num

class DataEntryApp:
    def __init__(self, master):
        self.master = master
        master.title('Tool nhập liệu Container |                                                                            Chế tạo bởi Minh Quang')
        # Powered by MinhQuang3tarots (hidden signature)
        self.master._minhquang3tarots = 'Powered by MinhQuang3tarots'
        self.entries = {}
        # Load sheet names
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE, read_only=True)
            self.sheet_names = [ws.title for ws in wb.worksheets if ws.sheet_state == 'visible']
            wb.close()
        else:
            self.sheet_names = []
        # Initialize ExcelManager here
        self.excel_mgr = ExcelManager(EXCEL_FILE)
        # Sheet selection
        tk.Label(master, text='Chọn sheet:').grid(row=0, column=0, padx=5, pady=3, sticky='e')
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(master, textvariable=self.sheet_var, values=self.sheet_names, state='readonly')
        self.sheet_combo.grid(row=0, column=1, padx=5, pady=3)
        if self.sheet_names:
            self.sheet_combo.current(0)
        # Mapping sheet name to default company
        self.sheet_to_cty = {name: name for name in self.sheet_names}
        # Entry fields
        for idx, field in enumerate(FIELDS):
            label = tk.Label(master, text=field)
            label.grid(row=idx+1, column=0, padx=5, pady=3, sticky='e')
            if field == 'Loại hình':
                entry = ttk.Combobox(master, values=['Nhập', 'Xuất'], state='readonly')
                entry.set('Xuất')
            elif field == 'CTY':
                entry = ttk.Combobox(master, values=self.sheet_names, state='readonly')
                # Set default value for CTY based on selected sheet
                if self.sheet_names:
                    entry.set(self.sheet_var.get() or self.sheet_names[0])
                self.cty_entry = entry
            elif field == 'NHÀ XE':
                entry = ttk.Combobox(master, values=['GH', 'HNP', 'DAP'], state='readonly')
                entry.set('GH')
            elif field == 'Số lượng':
                entry = tk.Spinbox(master, from_=1, to=1000, width=28)
                entry.delete(0, tk.END)
                entry.insert(0, '1')
            elif field == 'Kích cỡ':
                entry = ttk.Combobox(master, values=['40', '20', '45'], state='readonly')
                entry.set('40')
            else:
                entry = tk.Entry(master, width=30)
            entry.grid(row=idx+1, column=1, padx=5, pady=3)
            self.entries[field] = entry
        # Set default value for 'NGÀY LẤY' to today
        today_str = datetime.now().strftime('%d/%m/%Y')
        if 'NGÀY LẤY' in self.entries:
            self.entries['NGÀY LẤY'].insert(0, today_str)
        self.save_button = tk.Button(master, text='Lưu', command=self.save_data)
        self.save_button.grid(row=len(FIELDS)+1, column=0, pady=10, sticky='e')
        # Thêm nút Hoàn tác
        self.undo_button = tk.Button(master, text='Hoàn tác', command=self.undo_last_entry)
        self.undo_button.grid(row=len(FIELDS)+1, column=1, pady=10, sticky='w')
        self.last_entry_info = None
        # Bind sheet selection to update CTY
        self.sheet_combo.bind('<<ComboboxSelected>>', self.update_cty_by_sheet)
        # Load settings
        self.settings = load_settings()
        # Add settings button
        self.settings_button = tk.Button(master, text='Cài đặt vị trí nhập', command=self.open_settings_window)
        self.settings_button.grid(row=len(FIELDS)+2, column=0, columnspan=2, pady=5)
        # Add Excel preview frame
        self.preview_frame = tk.Frame(master)
        self.preview_frame.grid(row=0, column=3, rowspan=len(FIELDS)+3, padx=10, pady=5, sticky='n')
        tk.Label(self.preview_frame, text='Xem trước dữ liệu Excel').pack()
        self.preview_text = tk.Text(self.preview_frame, width=150, height=30, font=('Consolas', 10), wrap='none', xscrollcommand=None)
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # Thêm thanh scroll ngang
        xscroll = tk.Scrollbar(self.preview_frame, orient='horizontal', command=self.preview_text.xview)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.preview_text.config(xscrollcommand=xscroll.set)
        self.refresh_button = tk.Button(self.preview_frame, text='Làm mới', command=self.refresh_preview)
        self.refresh_button.pack(pady=5)
        # Bind sheet selection chỉ 1 lần
        self.sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_change)
        # Thêm nút Xoá dòng trước đó
        self.delete_prev_button = tk.Button(master, text='Xoá dòng trước đó', command=self.delete_previous_row)
        self.delete_prev_button.grid(row=len(FIELDS)+2, column=1, pady=5, sticky='w')
        self.refresh_preview()
        # Kéo dài cửa sổ chính
       

    def update_cty_by_sheet(self, event=None):
        sheet = self.sheet_var.get()
        if 'CTY' in self.entries and sheet:
            self.entries['CTY'].set(sheet)

    def open_settings_window(self):
        win = tk.Toplevel(self.master)
        win.title('Cài đặt vị trí nhập')
        tk.Label(win, text='Sheet').grid(row=0, column=0, padx=5, pady=3)
        tk.Label(win, text='Dòng bắt đầu').grid(row=0, column=1, padx=5, pady=3)
        tk.Label(win, text='Cột bắt đầu (A, B, ...)').grid(row=0, column=2, padx=5, pady=3)
        entries = {}
        for idx, sheet in enumerate(self.sheet_names):
            tk.Label(win, text=sheet).grid(row=idx+1, column=0, padx=5, pady=3)
            start_row = tk.Entry(win, width=5)
            start_col = tk.Entry(win, width=5)
            s = self.settings.get('sheets', {}).get(sheet, {"start_row": 2, "start_col": 1})
            start_row.insert(0, str(s.get('start_row', 2)))
            # Hiển thị cột dạng chữ cái
            col_letter = ''
            col_num = s.get('start_col', 1)
            while col_num > 0:
                col_num, rem = divmod(col_num - 1, 26)
                col_letter = chr(rem + ord('A')) + col_letter
            start_col.insert(0, col_letter or 'A')
            start_row.grid(row=idx+1, column=1, padx=5, pady=3)
            start_col.grid(row=idx+1, column=2, padx=5, pady=3)
            entries[sheet] = (start_row, start_col)
        def save_and_close():
            for sheet, (row_e, col_e) in entries.items():
                try:
                    row = int(row_e.get())
                    col = col_letter_to_index(col_e.get())
                    if row < 1: row = 1
                    if col < 1: col = 1
                    self.settings['sheets'][sheet] = {"start_row": row, "start_col": col}
                except Exception:
                    continue
            save_settings(self.settings)
            win.destroy()
        tk.Button(win, text='Lưu', command=save_and_close).grid(row=len(self.sheet_names)+1, column=0, columnspan=3, pady=5)

    def on_sheet_change(self, event=None):
        self.update_cty_by_sheet()
        self.refresh_preview()

    def refresh_preview(self):
        self.preview_text.delete('1.0', tk.END)
        sheet_name = self.sheet_var.get() or (self.sheet_names[0] if self.sheet_names else None)
        if not sheet_name:
            self.preview_text.insert(tk.END, 'Không tìm thấy file hoặc sheet!')
            return
        try:
            s = self.settings.get('sheets', {}).get(sheet_name, {"start_row": 2, "start_col": 1})
            start_row = s.get('start_row', 2)
            start_col = s.get('start_col', 1)
            num_fields = len(FIELDS)
            # Nếu có cột STT ở cột A (start_col == 2), cần ghi STT tự động
            stt_offset = 1 if start_col == 2 else 0
            true_start_col = start_col + stt_offset
            # Lấy thêm cột STT (A) nếu có
            preview_num_fields = num_fields + stt_offset
            preview_start_col = start_col - stt_offset if stt_offset else start_col
            cell_cache, row, min_row, max_row = self.excel_mgr.preview_rows(sheet_name, start_row, preview_start_col, preview_num_fields)
            col_widths = []
            preview_rows = list(range(min_row, max_row + 1))
            for idx, c in enumerate(range(preview_start_col, preview_start_col + preview_num_fields)):
                col_letter = get_column_letter(c)
                max_len = len(col_letter)
                for r in preview_rows:
                    val = cell_cache.get((r, c))
                    val_str = str(val) if val is not None else ''
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                col_widths.append(max(max_len, 5))
            header = []
            for idx, c in enumerate(range(preview_start_col, preview_start_col + preview_num_fields)):
                col_letter = ''
                col_num = c
                while col_num > 0:
                    col_num, rem = divmod(col_num - 1, 26)
                    col_letter = chr(rem + ord('A')) + col_letter
                if stt_offset and idx == 0:
                    header.append('STT'.center(col_widths[idx]))
                else:
                    header.append(col_letter.center(col_widths[idx]))
            header_line = '| ' + ' | '.join(header) + ' |\n'
            sep_line = '+-' + '-+-'.join(['-' * w for w in col_widths]) + '-+\n'
            lines = []
            lines.append('    ' + sep_line)
            lines.append('    ' + header_line)
            lines.append('    ' + sep_line)
            for r in preview_rows:
                row_cells = []
                for idx, c in enumerate(range(preview_start_col, preview_start_col + preview_num_fields)):
                    val = cell_cache.get((r, c))
                    val_str = str(val) if val is not None else ''
                    pad = col_widths[idx] - len(val_str)
                    left = pad // 2
                    right = pad - left
                    cell_content = ' ' * left + val_str + ' ' * right
                    row_cells.append(cell_content)
                line = '| ' + ' | '.join(row_cells) + ' |'
                if r == row:
                    lines.append(f'>>> {line}\n')
                else:
                    lines.append(f'   {line}\n')
                lines.append('    ' + sep_line)
            self.preview_text.insert(tk.END, ''.join(lines))
        except Exception as e:
            self.preview_text.insert(tk.END, f'Lỗi khi đọc file: {e}')

    def save_data(self):
        data = []
        for field in FIELDS:
            if field == 'Loại hình':
                value = self.entries[field].get()
                if value == 'Xuất':
                    data.append('X')
                elif value == 'Nhập':
                    data.append('N')
                else:
                    data.append(value)
            else:
                data.append(self.entries[field].get())
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            messagebox.showerror('Lỗi', 'Vui lòng chọn sheet!')
            return
        try:
            s = self.settings.get('sheets', {}).get(sheet_name, {"start_row": 2, "start_col": 1})
            start_row = s.get('start_row', 2)
            start_col = s.get('start_col', 1)
            num_fields = len(FIELDS)
            # Nếu có cột STT ở cột A (start_col == 2), cần ghi STT tự động
            stt_offset = 1 if start_col == 2 else 0
            true_start_col = start_col + stt_offset
            ws = self.excel_mgr.get_sheet(sheet_name)
            # Tìm dòng ghi dữ liệu tiếp theo
            row = self.excel_mgr.get_last_empty_row(sheet_name, start_row, start_col, num_fields)
            # Ghi STT nếu có cột STT
            if stt_offset:
                ws.range((row, 1)).value = row - start_row + 1  # STT bắt đầu từ 1 tại start_row
            # Ghi dữ liệu các trường còn lại
            for i, value in enumerate(data):
                ws.range((row, true_start_col + i)).value = value
            self.excel_mgr.save()
            self.excel_mgr.clear_last_empty_row_cache(sheet_name)
            self.last_entry_info = {
                'sheet': sheet_name,
                'row': row,
                'start_col': true_start_col,
                'num_fields': len(data),
                'values': data.copy()
            }
            messagebox.showinfo('Thành công', f'Đã lưu dữ liệu vào dòng {row}!')
            for entry in self.entries.values():
                entry.delete(0, tk.END)
            today_str = datetime.now().strftime('%d/%m/%Y')
            if 'NGÀY LẤY' in self.entries:
                self.entries['NGÀY LẤY'].insert(0, today_str)
            if 'Loại hình' in self.entries:
                self.entries['Loại hình'].set('Xuất')
            self.refresh_preview()
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể lưu dữ liệu: {e}')

    def undo_last_entry(self):
        if not self.last_entry_info:
            messagebox.showinfo('Hoàn tác', 'Không có thao tác nào để hoàn tác!')
            return
        try:
            sheet = self.last_entry_info['sheet']
            row = self.last_entry_info['row']
            start_col = self.last_entry_info['start_col']
            num_fields = self.last_entry_info['num_fields']
            values = self.last_entry_info['values']
            self.excel_mgr.undo_row(sheet, row, start_col, num_fields)
            self.excel_mgr.save()
            for idx, field in enumerate(FIELDS):
                entry = self.entries[field]
                entry.delete(0, tk.END)
                entry.insert(0, str(values[idx]))
                if isinstance(entry, ttk.Combobox):
                    entry.set(str(values[idx]))
            self.settings['sheets'][sheet]['start_row'] = row
            save_settings(self.settings)
            self.last_entry_info = None
            messagebox.showinfo('Hoàn tác', f'Đã hoàn tác dòng {row} trên sheet {sheet}! Dữ liệu đã được trả lại vào form và đã xoá khỏi Excel.')
            self.refresh_preview()
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể hoàn tác: {e}')

    def delete_previous_row(self):
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            messagebox.showerror('Lỗi', 'Vui lòng chọn sheet!')
            return
        s = self.settings.get('sheets', {}).get(sheet_name, {"start_row": 2, "start_col": 1})
        start_row = s.get('start_row', 2)
        start_col = s.get('start_col', 1)
        num_fields = len(FIELDS)
        # Nếu bảng có cột STT ở cột A, cần dịch start_col sang phải 1 cột
        stt_offset = 1 if start_col == 2 else 0  # Nếu start_col là 2 (B), thì offset 1 cho cột A là STT
        true_start_col = start_col + stt_offset
        # Xác định dòng ngay trước dòng bắt đầu nhập liệu (start_row - 1)
        prev_row = start_row - 1
        if prev_row < 1:
            messagebox.showinfo('Thông báo', 'Không còn dòng nào để xoá!')
            return
        try:
            ws = self.excel_mgr.get_sheet(sheet_name)
            is_empty = all(ws.range((prev_row, true_start_col + i)).value in (None, '') for i in range(num_fields))
            if is_empty:
                messagebox.showinfo('Thông báo', f'Dòng {prev_row} đã trống!')
                return
            for i in range(num_fields):
                ws.range((prev_row, true_start_col + i)).value = None
            self.excel_mgr.save()
            self.excel_mgr.clear_last_empty_row_cache(sheet_name)
            self.settings['sheets'][sheet_name]['start_row'] = prev_row
            save_settings(self.settings)
            messagebox.showinfo('Thành công', f'Đã xoá dữ liệu dòng {prev_row} trên sheet {sheet_name}! Vị trí nhập liệu sẽ bắt đầu lại từ dòng này.')
            self.refresh_preview()
        except Exception as e:
            messagebox.showerror('Lỗi', f'Không thể xoá dòng {prev_row}: {e}')

    def on_close(self):
        self.excel_mgr.close()
        self.master.destroy()

def main():
    root = tk.Tk()
    app = DataEntryApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()

if __name__ == '__main__':
    main()
